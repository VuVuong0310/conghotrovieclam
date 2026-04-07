import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ============================================================
# SHEET 1: BIEU MAU CHAM DIEM
# ============================================================
ws = wb.active
ws.title = "J2EE"

# Styles
bold = Font(bold=True, size=11)
bold_big = Font(bold=True, size=13)
header_font = Font(bold=True, size=11, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
basic_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
advanced_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
deploy_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
total_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
thin_border = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
wrap = Alignment(wrap_text=True, vertical='center')
center = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Column widths
ws.column_dimensions['A'].width = 15
ws.column_dimensions['B'].width = 20
ws.column_dimensions['C'].width = 20
ws.column_dimensions['D'].width = 35
ws.column_dimensions['E'].width = 8
ws.column_dimensions['F'].width = 55
ws.column_dimensions['G'].width = 8

# === HEADER ===
ws['A1'] = 'Thông tin Nhóm'
ws['A1'].font = bold_big

ws['A2'] = 'Nhóm'
ws['A2'].font = bold

ws['A3'] = 'Tên đồ án'
ws['A3'].font = bold
ws['B3'] = 'Cổng Hỗ Trợ Việc Làm (Job Portal)'

ws['A4'] = 'Công nghệ Backend(API)'
ws['A4'].font = bold
ws['B4'] = 'Spring Boot 3.2.3 + Java 19 + Spring Security + JWT'

ws['A5'] = 'CSDL sử dụng'
ws['A5'].font = bold
ws['B5'] = 'MySQL 8.0'

# === TABLE HEADER (row 7) ===
headers = ['MSSV', 'Họ tên SV', 'Nhóm chức năng', 'Chức năng', 'Có làm', 'Mô tả', 'Điểm']
for col, h in enumerate(headers, 1):
    cell = ws.cell(row=7, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# === CHUC NANG CO BAN (5 diem) ===
row = 8
ws.cell(row=row, column=3, value='Cơ bản').font = bold
ws.cell(row=row, column=3).fill = basic_fill

basic_features = [
    ('Đăng ký / Đăng nhập / Phân quyền', 'Có', 'Register (CANDIDATE/EMPLOYER), Login JWT, 3 roles (CANDIDATE, EMPLOYER, ADMIN)'),
    ('CRUD ít nhất 5 bảng', 'Có', 'JobPost, Education, Experience, Project, JobCategory, User, Notification (7 bảng full CRUD)'),
    ('Đăng nhập bằng Google', 'Có', 'Google OAuth2 login tích hợp'),
    ('CRUD không ảnh hưởng bảng khác', 'Có', 'Education, Experience, Project CRUD độc lập'),
    ('Bật / tắt / block trạng thái', 'Có', 'Lock/unlock user, toggle job active, approve/reject job, 6 trạng thái application'),
    ('Tìm kiếm / sắp xếp', 'Có', 'Tìm kiếm theo keyword, location, salary, loại hình + sắp xếp nhiều trường + phân trang'),
    ('Upload file (Resume + Photo)', 'Có', 'Resume PDF (max 5MB) + Profile photo (JPG/PNG)'),
    ('Email thông báo', 'Có', 'Email xác nhận ứng tuyển, phỏng vấn, từ chối, chấp nhận, reset password'),
    ('Gợi ý việc làm phù hợp', 'Có', 'Hệ thống gợi ý việc làm dựa trên kỹ năng, vị trí, địa điểm, mức lương'),
    ('Hệ thống thông báo', 'Có', 'Thông báo trong ứng dụng, đếm chưa đọc, đánh dấu đã đọc'),
]

for i, (func, co, mota) in enumerate(basic_features):
    r = row + i
    ws.cell(row=r, column=4, value=func).border = thin_border
    ws.cell(row=r, column=4).alignment = wrap
    ws.cell(row=r, column=5, value=co).border = thin_border
    ws.cell(row=r, column=5).alignment = center
    ws.cell(row=r, column=6, value=mota).border = thin_border
    ws.cell(row=r, column=6).alignment = wrap
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=3).fill = basic_fill
    ws.cell(row=r, column=7).border = thin_border

# Diem co ban
ws.cell(row=row, column=7, value='5').font = bold
ws.cell(row=row, column=7).alignment = center
ws.cell(row=row, column=7).border = thin_border

row = row + len(basic_features) + 1

# === DANH SACH CAC BANG DA CRUD ===
ws.cell(row=row, column=3, value='Danh sách các bảng đã CRUD').font = bold
ws.cell(row=row, column=3).fill = basic_fill

crud_tables = [
    ('JobPost (Tin tuyển dụng)', 'CRUD', 'C: POST /api/jobs, R: GET, U: PUT, D: DELETE'),
    ('Education (Học vấn)', 'CRUD', 'Full CRUD qua /api/candidates/{id}/educations'),
    ('Experience (Kinh nghiệm)', 'CRUD', 'Full CRUD qua /api/candidates/{id}/experiences'),
    ('Project (Dự án)', 'CRUD', 'Full CRUD qua /api/candidates/{id}/projects'),
    ('JobCategory (Danh mục)', 'CRUD', 'C: POST, R: GET, U: PUT, D: DELETE (Admin)'),
    ('User (Người dùng)', 'CRUD', 'C: Register, R: GET, U: toggle-lock, D: DELETE (Admin)'),
    ('Notification (Thông báo)', 'CRD', 'C: auto, R: GET, U: mark-read, D: DELETE'),
    ('CandidateProfile', 'CRU', 'C: auto, R: GET, U: PUT (+ upload resume/photo)'),
    ('EmployerProfile', 'RU', 'R: GET, U: PUT company-profile'),
    ('JobApplication (Ứng tuyển)', 'CRU', 'C: POST apply, R: GET, U: update status'),
]

for i, (table, crud, mota) in enumerate(crud_tables):
    r = row + i
    ws.cell(row=r, column=4, value=table).border = thin_border
    ws.cell(row=r, column=4).alignment = wrap
    ws.cell(row=r, column=5, value=crud).border = thin_border
    ws.cell(row=r, column=5).alignment = center
    ws.cell(row=r, column=6, value=mota).border = thin_border
    ws.cell(row=r, column=6).alignment = wrap
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=3).fill = basic_fill
    ws.cell(row=r, column=7).border = thin_border

row = row + len(crud_tables) + 1

# === DEPLOY (2 diem) ===
ws.cell(row=row, column=3, value='Deploy Web/app').font = bold
ws.cell(row=row, column=3).fill = deploy_fill
ws.cell(row=row, column=4, value='https://conghotrovieclam.online').border = thin_border
ws.cell(row=row, column=4).alignment = wrap
ws.cell(row=row, column=5, value='Có').border = thin_border
ws.cell(row=row, column=5).alignment = center
ws.cell(row=row, column=6, value='Domain riêng có phí + HTTPS/SSL + Docker Compose').border = thin_border
ws.cell(row=row, column=6).alignment = wrap
ws.cell(row=row, column=7, value='2').font = bold
ws.cell(row=row, column=7).alignment = center
ws.cell(row=row, column=7).border = thin_border
ws.cell(row=row, column=3).border = thin_border

row += 2

# === CHUC NANG NANG CAO (max 3 diem) ===
ws.cell(row=row, column=3, value='Chức năng nâng cao').font = bold
ws.cell(row=row, column=3).fill = advanced_fill
ws.cell(row=row, column=6, value='0.5 cho mỗi chức năng (tối đa 3 điểm)').border = thin_border
ws.cell(row=row, column=6).alignment = wrap
ws.cell(row=row, column=7, value='3').font = bold
ws.cell(row=row, column=7).alignment = center
ws.cell(row=row, column=7).border = thin_border
ws.cell(row=row, column=3).border = thin_border

# Advanced features - CHI GHI CHUC NANG, KHONG GHI CONG NGHE
advanced_features = [
    ('Đăng nhập bằng tài khoản Google', 'Có', 'Cho phép người dùng đăng nhập bằng tài khoản Google'),
    ('Gửi email thông báo tự động', 'Có', 'Tự động gửi email khi ứng tuyển, phỏng vấn, từ chối, reset mật khẩu'),
    ('Upload và quản lý hồ sơ (Resume + Ảnh)', 'Có', 'Upload file hồ sơ PDF và ảnh đại diện, xem và tải về'),
    ('Gợi ý việc làm phù hợp', 'Có', 'Hệ thống gợi ý việc làm dựa trên kỹ năng, chức danh, địa điểm, mức lương'),
    ('Xác thực và phân quyền người dùng', 'Có', 'Đăng nhập an toàn, phân quyền theo vai trò, bảo vệ API'),
    ('Đóng gói và triển khai tự động', 'Có', 'Đóng gói ứng dụng thành container, triển khai nhiều dịch vụ cùng lúc'),
    ('Bảo mật HTTPS và chứng chỉ SSL', 'Có', 'Kết nối an toàn qua HTTPS với chứng chỉ SSL tự động'),
    ('Tìm kiếm nâng cao đa tiêu chí', 'Có', 'Tìm kiếm kết hợp nhiều điều kiện: từ khóa, địa điểm, mức lương, loại hình'),
    ('Xuất CV nhiều mẫu (8 mẫu)', 'Có', 'Tạo CV từ thông tin hồ sơ với 8 mẫu giao diện khác nhau'),
    ('Hệ thống thông báo trong ứng dụng', 'Có', 'Thông báo tự động khi có sự kiện, đếm chưa đọc, đánh dấu đã đọc'),
    ('Quản lý trạng thái ứng tuyển', 'Có', 'Theo dõi 6 trạng thái: Mới, Xem, Phỏng vấn, Chấp nhận, Từ chối, Rút'),
    ('Danh sách công ty và trang chi tiết', 'Có', 'Xem danh sách công ty, thông tin chi tiết, các việc làm của công ty'),
]

for i, (func, co, mota) in enumerate(advanced_features):
    r = row + 1 + i
    ws.cell(row=r, column=4, value=func).border = thin_border
    ws.cell(row=r, column=4).alignment = wrap
    ws.cell(row=r, column=5, value=co).border = thin_border
    ws.cell(row=r, column=5).alignment = center
    ws.cell(row=r, column=6, value=mota).border = thin_border
    ws.cell(row=r, column=6).alignment = wrap
    ws.cell(row=r, column=3).border = thin_border
    ws.cell(row=r, column=3).fill = advanced_fill
    ws.cell(row=r, column=7, value='0.5').border = thin_border
    ws.cell(row=r, column=7).alignment = center

row = row + 1 + len(advanced_features) + 1

# === TONG DIEM ===
ws.cell(row=row, column=1, value='TỔNG ĐIỂM').font = Font(bold=True, size=14)
ws.cell(row=row, column=1).fill = total_fill
for c in range(1, 8):
    ws.cell(row=row, column=c).fill = total_fill
    ws.cell(row=row, column=c).border = thin_border
ws.cell(row=row, column=7, value=10).font = Font(bold=True, size=14)
ws.cell(row=row, column=7).alignment = center

row += 2

# === CHI TIET TONG DIEM ===
ws.cell(row=row, column=3, value='Chi tiết tổng điểm:').font = bold
row += 1
details = [
    ('Chức năng cơ bản', '5'),
    ('Deploy có phí (domain riêng)', '2'),
    ('Chức năng nâng cao (12 x 0.5, tối đa 3)', '3'),
    ('TỔNG', '10'),
]
for label, pts in details:
    ws.cell(row=row, column=3, value=label).border = thin_border
    ws.cell(row=row, column=3).alignment = wrap
    ws.cell(row=row, column=7, value=pts).border = thin_border
    ws.cell(row=row, column=7).alignment = center
    if label == 'TỔNG':
        ws.cell(row=row, column=3).font = bold
        ws.cell(row=row, column=7).font = bold
        ws.cell(row=row, column=3).fill = total_fill
        ws.cell(row=row, column=7).fill = total_fill
    row += 1

# ============================================================
# SHEET 2: PHAN CONG CONG VIEC
# ============================================================
ws2 = wb.create_sheet("Phân công")

# Styles for sheet 2
sv_fill_vuong = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")   # xanh nhat
sv_fill_khiem = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")    # xanh la
sv_fill_nhat = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")     # cam
sv_fill_phung = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")    # vang
sv_fill_cong = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")     # tim
sv_fill_thinh = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")    # hong
sv_fill_tri = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")      # xanh dam
sv_fill_dang = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")     # xam

ws2.column_dimensions['A'].width = 16
ws2.column_dimensions['B'].width = 22
ws2.column_dimensions['C'].width = 10
ws2.column_dimensions['D'].width = 18
ws2.column_dimensions['E'].width = 45
ws2.column_dimensions['F'].width = 50

# Title
ws2['A1'] = 'BẢNG PHÂN CÔNG CÔNG VIỆC - JOB PORTAL'
ws2['A1'].font = Font(bold=True, size=14)

# Header row 3
pc_headers = ['MSSV', 'Họ tên SV', 'Điểm mong muốn', 'Vai trò', 'Chức năng phụ trách', 'Chi tiết công việc']
for col, h in enumerate(pc_headers, 1):
    cell = ws2.cell(row=3, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border

# =============================================
# PHAN CONG CHI TIET
# =============================================
# Moi nguoi co nhieu dong, moi dong la 1 chuc nang
students = [
    # (MSSV, Tên, Điểm, Vai trò, fill, [(chức_năng, chi_tiết), ...])
    ('2280603777', 'Vũ Văn Vương', 9, 'Code chính', sv_fill_vuong, [
        ('Thiết kế kiến trúc hệ thống', 'Thiết kế toàn bộ kiến trúc Backend (Spring Boot) và Frontend (React), cấu hình dự án'),
        ('Đăng ký / Đăng nhập / Phân quyền', 'Xây dựng hệ thống xác thực JWT, phân quyền 3 role (CANDIDATE, EMPLOYER, ADMIN), Spring Security'),
        ('Đăng nhập bằng Google', 'Tích hợp Google OAuth2 authorization code flow, xử lý callback, tạo JWT'),
        ('Gợi ý việc làm phù hợp', 'Xây dựng thuật toán matching: skills (40đ), title (30đ), location (15đ), salary (15đ)'),
        ('Xuất CV 8 mẫu giao diện', 'Lập trình 8 template CV (classic, modern, minimal, elegant, creative, executive, developer, academic)'),
        ('Deploy (Docker + Cloud)', 'Cấu hình Docker Compose, Cloudflare Tunnel, tên miền có phí, SSL/HTTPS'),
        ('Tìm kiếm nâng cao', 'Xây dựng Advanced Search đa tiêu chí (keyword, location, salary, type) + pagination + sort'),
        ('Bảo mật API', 'Cấu hình Spring Security, @PreAuthorize, bảo vệ endpoint theo role'),
    ]),
    ('1228060149', 'Lê Trần Gia Khiêm', 8.5, 'Merge code / Quản lý nhóm / Code chính', sv_fill_khiem, [
        ('Quản lý source code', 'Quản lý nhánh Git, merge code các thành viên, giải quyết conflict, review code'),
        ('CRUD JobPost (Tin tuyển dụng)', 'Backend + Frontend: Tạo, xem, sửa, xóa tin tuyển dụng (Employer)'),
        ('Quản lý ứng tuyển (JobApplication)', 'Backend + Frontend: Ứng viên nộp hồ sơ, Employer duyệt 6 trạng thái'),
        ('Email thông báo', 'Cấu hình Gmail SMTP, gửi email xác nhận ứng tuyển, phỏng vấn, từ chối, reset password'),
        ('Bật / tắt / block trạng thái', 'Lock/unlock user, toggle job active, approve/reject job'),
        ('Employer Dashboard', 'Frontend: Trang quản lý của nhà tuyển dụng, danh sách tin, danh sách ứng tuyển'),
    ]),
    ('2280600674', 'Trần Hoàng Thế Đăng', 8.5, 'Không tham gia', sv_fill_dang, [
        ('(Không tham gia)', 'Không đóng góp code'),
    ]),
    ('2280602212', 'Phạm Minh Nhật', 8, 'Code cơ bản', sv_fill_nhat, [
        ('CRUD Education (Học vấn)', 'Backend + Frontend: Thêm, sửa, xóa thông tin học vấn của ứng viên'),
        ('CRUD Experience (Kinh nghiệm)', 'Backend + Frontend: Thêm, sửa, xóa kinh nghiệm làm việc của ứng viên'),
        ('Trang Candidate Profile', 'Frontend: Hiển thị và chỉnh sửa thông tin cá nhân ứng viên'),
    ]),
    ('2280602491', 'Lâm Minh Phụng', 8, 'Code cơ bản', sv_fill_phung, [
        ('CRUD Project (Dự án)', 'Backend + Frontend: Thêm, sửa, xóa dự án cá nhân của ứng viên'),
        ('CRUD JobCategory (Danh mục)', 'Backend + Frontend: Quản lý danh mục ngành nghề (Admin)'),
        ('Trang Đăng ký + Đăng nhập', 'Frontend: Form register (Candidate/Employer), form login, forgot password'),
    ]),
    ('2280600329', 'Trần Thành Công', 8, 'Code cơ bản', sv_fill_cong, [
        ('Upload file (Resume + Photo)', 'Backend + Frontend: Upload PDF resume (max 5MB), profile photo (JPG/PNG), download'),
        ('Trang My Applications', 'Frontend: Danh sách việc đã ứng tuyển, trạng thái, rút hồ sơ'),
        ('Trang About / Giới thiệu', 'Frontend: Trang giới thiệu hệ thống'),
    ]),
    ('2280603108', 'Phạm Lê Gia Thịnh', 8, 'Code cơ bản', sv_fill_thinh, [
        ('Hệ thống thông báo (Notification)', 'Backend + Frontend: Thông báo tự động, đếm chưa đọc, đánh dấu đã đọc, xóa'),
        ('CRUD Notification', 'Backend: API tạo/đọc/xóa thông báo, Frontend: component thông báo'),
        ('Admin Dashboard', 'Frontend: Trang quản lý của Admin (thống kê, quản lý user)'),
    ]),
    ('2280603381', 'Vũ Mạnh Trí', 8, 'Code cơ bản', sv_fill_tri, [
        ('Tìm kiếm / sắp xếp cơ bản', 'Frontend: Trang tìm việc với filter, sort, pagination'),
        ('Danh sách công ty', 'Backend + Frontend: Hiển thị danh sách công ty, trang chi tiết, việc làm của công ty'),
        ('Trang Job Detail', 'Frontend: Trang chi tiết tin tuyển dụng, nút ứng tuyển'),
    ]),
]

row2 = 4
for mssv, ten, diem, vaitro, fill, tasks in students:
    start_row = row2
    for i, (chucnang, chitiet) in enumerate(tasks):
        r = row2 + i
        if i == 0:
            ws2.cell(row=r, column=1, value=mssv).border = thin_border
            ws2.cell(row=r, column=1).alignment = center
            ws2.cell(row=r, column=2, value=ten).border = thin_border
            ws2.cell(row=r, column=2).alignment = wrap
            ws2.cell(row=r, column=3, value=diem).border = thin_border
            ws2.cell(row=r, column=3).alignment = center
            ws2.cell(row=r, column=4, value=vaitro).border = thin_border
            ws2.cell(row=r, column=4).alignment = wrap
        else:
            for c in range(1, 5):
                ws2.cell(row=r, column=c).border = thin_border

        ws2.cell(row=r, column=5, value=chucnang).border = thin_border
        ws2.cell(row=r, column=5).alignment = wrap
        ws2.cell(row=r, column=6, value=chitiet).border = thin_border
        ws2.cell(row=r, column=6).alignment = wrap

        # Apply fill color for entire row
        for c in range(1, 7):
            ws2.cell(row=r, column=c).fill = fill

    # Merge cells for MSSV, Ten, Diem, Vaitro
    end_row = row2 + len(tasks) - 1
    if len(tasks) > 1:
        ws2.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
        ws2.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        ws2.merge_cells(start_row=start_row, start_column=3, end_row=end_row, end_column=3)
        ws2.merge_cells(start_row=start_row, start_column=4, end_row=end_row, end_column=4)

    row2 = end_row + 1

# === THONG KE PHAN CONG ===
row2 += 1
ws2.cell(row=row2, column=1, value='THỐNG KÊ PHÂN CÔNG').font = Font(bold=True, size=12)
row2 += 1

stat_headers = ['MSSV', 'Họ tên', 'Vai trò', 'Số chức năng', 'Điểm mong muốn', 'Ghi chú']
for col, h in enumerate(stat_headers, 1):
    cell = ws2.cell(row=row2, column=col, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center
    cell.border = thin_border
row2 += 1

notes = {
    'Vũ Văn Vương': 'Code chính: kiến trúc + nâng cao + deploy',
    'Lê Trần Gia Khiêm': 'Merge code, quản lý nhóm + code chính',
    'Trần Hoàng Thế Đăng': 'Không tham gia',
    'Phạm Minh Nhật': 'CRUD cơ bản: Education, Experience',
    'Lâm Minh Phụng': 'CRUD cơ bản: Project, Category, Login page',
    'Trần Thành Công': 'Upload file, My Applications',
    'Phạm Lê Gia Thịnh': 'Notification, Admin Dashboard',
    'Vũ Mạnh Trí': 'Tìm kiếm, Company list, Job detail',
}

for mssv, ten, diem, vaitro, fill, tasks in students:
    ws2.cell(row=row2, column=1, value=mssv).border = thin_border
    ws2.cell(row=row2, column=1).alignment = center
    ws2.cell(row=row2, column=2, value=ten).border = thin_border
    ws2.cell(row=row2, column=2).alignment = wrap
    ws2.cell(row=row2, column=3, value=vaitro).border = thin_border
    ws2.cell(row=row2, column=3).alignment = wrap
    ws2.cell(row=row2, column=4, value=len(tasks)).border = thin_border
    ws2.cell(row=row2, column=4).alignment = center
    ws2.cell(row=row2, column=5, value=diem).border = thin_border
    ws2.cell(row=row2, column=5).alignment = center
    ws2.cell(row=row2, column=6, value=notes.get(ten, '')).border = thin_border
    ws2.cell(row=row2, column=6).alignment = wrap
    for c in range(1, 7):
        ws2.cell(row=row2, column=c).fill = fill
    row2 += 1

# Save
output = 'Bieu_mau_cham_diem_JobPortal_v3.xlsx'
wb.save(output)
print(f'Saved: {output}')
print()
print('=== PHAN CONG ===')
for mssv, ten, diem, vaitro, fill, tasks in students:
    print(f'{ten} ({diem}d) - {vaitro}: {len(tasks)} chuc nang')
    for cn, ct in tasks:
        print(f'  - {cn}')
print()
print('Tong: 8 thanh vien, 2 sheets (J2EE + Phan cong)')
